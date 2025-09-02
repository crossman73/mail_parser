"""
통합 타임라인 생성기 - 이메일과 추가 증거를 시간순으로 통합
Integrated Timeline Generator for Court Submission

이메일과 추가 증거 자료를 발생 날짜 기준으로 통합하여
법원 제출용 시간순 타임라인을 생성합니다.
"""

import json
import os
import shutil
import zipfile
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Union


class IntegratedTimelineGenerator:
    """통합 타임라인 생성기"""

    def __init__(self, base_dir: str = "processed_emails"):
        self.base_dir = Path(base_dir)
        self.timeline_dir = self.base_dir / "06_통합타임라인"
        self.timeline_dir.mkdir(exist_ok=True)

    def generate_integrated_timeline(self) -> Dict:
        """이메일과 추가 증거를 통합한 타임라인 생성"""
        print("📅 통합 타임라인 생성을 시작합니다...")
        print("=" * 60)

        timeline_items = []

        # 1. 이메일 데이터 수집
        email_items = self._collect_email_timeline_data()
        timeline_items.extend(email_items)
        print(f"  📧 이메일 항목: {len(email_items)}개")

        # 2. 추가 증거 데이터 수집
        additional_items = self._collect_additional_evidence_data()
        timeline_items.extend(additional_items)
        print(f"  📎 추가 증거 항목: {len(additional_items)}개")

        # 3. 시간순 정렬
        timeline_items.sort(key=lambda x: x['event_date'])
        print(f"  📋 총 타임라인 항목: {len(timeline_items)}개")

        # 4. 타임라인 보고서 생성
        timeline_report = self._generate_timeline_report(timeline_items)

        # 5. 법원 제출용 패키지 생성
        self._create_court_submission_package(timeline_items)

        # 6. Excel 파일 생성
        excel_path = self._generate_excel_timeline(timeline_items)

        print("✅ 통합 타임라인 생성 완료!")
        return {
            'total_items': len(timeline_items),
            'email_items': len(email_items),
            'additional_items': len(additional_items),
            'timeline_items': timeline_items,
            'excel_path': str(excel_path),
            'timeline_report': timeline_report
        }

    def _collect_email_timeline_data(self) -> List[Dict]:
        """이메일 타임라인 데이터 수집"""
        email_items = []

        # 처리된 이메일 디렉토리 스캔
        for email_dir in self.base_dir.glob("[*]_*"):
            if not email_dir.is_dir():
                continue

            try:
                # 이메일 메타데이터 찾기
                metadata_files = list(email_dir.glob("*metadata*.json"))
                if not metadata_files:
                    continue

                with open(metadata_files[0], 'r', encoding='utf-8') as f:
                    metadata = json.load(f)

                # 이메일 발송 날짜 파싱
                email_date = self._parse_email_date(metadata.get('date', ''))
                if not email_date:
                    continue

                # 첨부파일 정보 수집
                attachments = []
                attachments_dir = email_dir / "첨부파일"
                if attachments_dir.exists():
                    for attachment_file in attachments_dir.iterdir():
                        if attachment_file.is_file():
                            attachments.append({
                                'name': attachment_file.name,
                                'size': attachment_file.stat().st_size,
                                'path': str(attachment_file.relative_to(self.base_dir))
                            })

                # 증거 파일 경로
                evidence_files = []
                pdf_files = list(email_dir.glob("*.pdf"))
                html_files = list(email_dir.glob("*.html"))

                for pdf_file in pdf_files:
                    evidence_files.append({
                        'type': 'PDF',
                        'name': pdf_file.name,
                        'path': str(pdf_file.relative_to(self.base_dir))
                    })

                for html_file in html_files:
                    evidence_files.append({
                        'type': 'HTML',
                        'name': html_file.name,
                        'path': str(html_file.relative_to(self.base_dir))
                    })

                email_item = {
                    'type': 'EMAIL',
                    'event_date': email_date,
                    'event_time': email_date.strftime('%H:%M:%S') if hasattr(email_date, 'hour') else '00:00:00',
                    'title': metadata.get('subject', '제목 없음'),
                    'description': f"발신자: {metadata.get('from', 'N/A')}\n수신자: {metadata.get('to', 'N/A')}",
                    'participants': {
                        'from': metadata.get('from', 'N/A'),
                        'to': metadata.get('to', 'N/A'),
                        'cc': metadata.get('cc', ''),
                        'bcc': metadata.get('bcc', '')
                    },
                    'attachments': attachments,
                    'evidence_files': evidence_files,
                    'folder_name': email_dir.name,
                    'message_id': metadata.get('message_id', ''),
                    'importance': self._determine_email_importance(metadata, attachments)
                }

                email_items.append(email_item)

            except Exception as e:
                print(f"  ⚠️ 이메일 데이터 수집 오류 ({email_dir.name}): {str(e)}")
                continue

        return email_items

    def _collect_additional_evidence_data(self) -> List[Dict]:
        """추가 증거 타임라인 데이터 수집"""
        additional_items = []

        additional_evidence_dir = self.base_dir / "05_추가증거"
        if not additional_evidence_dir.exists():
            return additional_items

        metadata_file = additional_evidence_dir / "추가증거_목록.json"
        if not metadata_file.exists():
            return additional_items

        try:
            with open(metadata_file, 'r', encoding='utf-8') as f:
                metadata = json.load(f)

            for file_id, file_info in metadata.get('files', {}).items():
                # 증거 발생 날짜 파싱
                evidence_date = self._parse_evidence_date(
                    file_info.get('evidence_date', ''))
                if not evidence_date:
                    continue

                # 실제 파일 경로 확인
                file_path = self.base_dir / file_info['current_path']
                if not file_path.exists():
                    print(f"  ⚠️ 추가 증거 파일 누락: {file_info['title']}")
                    continue

                additional_item = {
                    'type': 'ADDITIONAL_EVIDENCE',
                    'event_date': evidence_date,
                    'event_time': '00:00:00',  # 추가 증거는 보통 정확한 시간이 없음
                    'title': file_info['title'],
                    'description': file_info.get('description', ''),
                    'category': file_info.get('category_name', '기타'),
                    'evidence_number': file_info.get('evidence_number', ''),
                    'party': file_info.get('party', '갑'),
                    'file_info': {
                        'name': file_path.name,
                        'size': file_info['file_size'],
                        'type': file_info.get('mime_type', ''),
                        'path': file_info['current_path'],
                        'hash': file_info['file_hash']
                    },
                    'related_emails': file_info.get('related_email_ids', []),
                    'importance': self._determine_evidence_importance(file_info)
                }

                additional_items.append(additional_item)

        except Exception as e:
            print(f"  ⚠️ 추가 증거 데이터 수집 오류: {str(e)}")

        return additional_items

    def _parse_email_date(self, date_str: str) -> Optional[datetime]:
        """이메일 날짜 문자열 파싱"""
        if not date_str:
            return None

        # 다양한 날짜 형식 지원
        date_formats = [
            '%a, %d %b %Y %H:%M:%S %z',
            '%a, %d %b %Y %H:%M:%S',
            '%Y-%m-%d %H:%M:%S',
            '%Y-%m-%d',
            '%d/%m/%Y %H:%M:%S',
            '%d/%m/%Y'
        ]

        for fmt in date_formats:
            try:
                return datetime.strptime(date_str.strip(), fmt)
            except ValueError:
                continue

        # RFC 2822 형식 처리
        try:
            from email.utils import parsedate_to_datetime
            return parsedate_to_datetime(date_str)
        except:
            pass

        return None

    def _parse_evidence_date(self, date_str: str) -> Optional[date]:
        """증거 날짜 문자열 파싱"""
        if not date_str:
            return None

        date_formats = ['%Y-%m-%d', '%d/%m/%Y', '%Y.%m.%d', '%Y년 %m월 %d일']

        for fmt in date_formats:
            try:
                parsed_date = datetime.strptime(date_str.strip(), fmt)
                return parsed_date.date()
            except ValueError:
                continue

        return None

    def _determine_email_importance(self, metadata: Dict, attachments: List) -> str:
        """이메일 중요도 판단"""
        # 첨부파일이 많으면 중요
        if len(attachments) >= 3:
            return 'HIGH'

        # 제목에 중요 키워드가 있으면
        subject = metadata.get('subject', '').lower()
        important_keywords = ['계약', '합의', '중요', '긴급', '법적', '소송', '고발', '신고']
        if any(keyword in subject for keyword in important_keywords):
            return 'HIGH'

        # 첨부파일이 있으면 보통
        if len(attachments) > 0:
            return 'MEDIUM'

        return 'LOW'

    def _determine_evidence_importance(self, file_info: Dict) -> str:
        """추가 증거 중요도 판단"""
        # 계약서, 법적 문서는 높음
        title = file_info.get('title', '').lower()
        description = file_info.get('description', '').lower()

        high_keywords = ['계약서', '합의서', '법적', '소송', '판결', '결정', '명령']
        if any(keyword in title or keyword in description for keyword in high_keywords):
            return 'HIGH'

        # 파일 크기가 큰 경우 (문서나 영상 등)
        if file_info.get('file_size', 0) > 10 * 1024 * 1024:  # 10MB 이상
            return 'MEDIUM'

        return 'MEDIUM'  # 추가 증거는 기본적으로 중요

    def _generate_timeline_report(self, timeline_items: List[Dict]) -> Dict:
        """타임라인 보고서 생성"""
        if not timeline_items:
            return {'summary': '타임라인 항목이 없습니다.'}

        # 기간 계산
        start_date = min(item['event_date'] for item in timeline_items)
        end_date = max(item['event_date'] for item in timeline_items)

        if isinstance(start_date, datetime):
            start_date = start_date.date()
        if isinstance(end_date, datetime):
            end_date = end_date.date()

        duration = (end_date - start_date).days

        # 유형별 통계
        email_count = len(
            [item for item in timeline_items if item['type'] == 'EMAIL'])
        evidence_count = len(
            [item for item in timeline_items if item['type'] == 'ADDITIONAL_EVIDENCE'])

        # 중요도별 통계
        high_importance = len(
            [item for item in timeline_items if item['importance'] == 'HIGH'])
        medium_importance = len(
            [item for item in timeline_items if item['importance'] == 'MEDIUM'])
        low_importance = len(
            [item for item in timeline_items if item['importance'] == 'LOW'])

        report = {
            'summary': f"{start_date}부터 {end_date}까지 {duration}일간의 타임라인",
            'period': {
                'start_date': str(start_date),
                'end_date': str(end_date),
                'duration_days': duration
            },
            'statistics': {
                'total_items': len(timeline_items),
                'email_count': email_count,
                'evidence_count': evidence_count,
                'high_importance': high_importance,
                'medium_importance': medium_importance,
                'low_importance': low_importance
            }
        }

        return report

    def _create_court_submission_package(self, timeline_items: List[Dict]):
        """법원 제출용 패키지 생성 (실제 파일 포함)"""
        print("📦 법원 제출용 패키지 생성 중...")

        # 패키지 디렉토리 생성
        package_dir = self.timeline_dir / \
            f"법원제출용_통합증거_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        package_dir.mkdir(exist_ok=True)

        # 타임라인 순서로 파일 복사
        for i, item in enumerate(timeline_items, 1):
            item_dir = package_dir / \
                f"{i:03d}_{item['event_date']}_{item['type']}"
            item_dir.mkdir(exist_ok=True)

            # 설명 파일 생성
            info_file = item_dir / "증거정보.txt"
            with open(info_file, 'w', encoding='utf-8') as f:
                f.write(f"증거 순번: {i}\n")
                f.write(f"발생 날짜: {item['event_date']}\n")
                f.write(f"증거 유형: {item['type']}\n")
                f.write(f"제목: {item['title']}\n")
                f.write(f"설명: {item['description']}\n")
                f.write(f"중요도: {item['importance']}\n\n")

                if item['type'] == 'EMAIL':
                    f.write("=== 이메일 정보 ===\n")
                    f.write(f"발신자: {item['participants']['from']}\n")
                    f.write(f"수신자: {item['participants']['to']}\n")
                    if item['participants']['cc']:
                        f.write(f"참조: {item['participants']['cc']}\n")
                    f.write(f"메시지 ID: {item['message_id']}\n")

                    # 첨부파일 정보
                    if item['attachments']:
                        f.write(f"\n첨부파일 ({len(item['attachments'])}개):\n")
                        for att in item['attachments']:
                            f.write(
                                f"  - {att['name']} ({att['size']} bytes)\n")

                    # 증거 파일 복사
                    for evidence_file in item['evidence_files']:
                        source_path = self.base_dir / evidence_file['path']
                        if source_path.exists():
                            dest_path = item_dir / evidence_file['name']
                            shutil.copy2(source_path, dest_path)

                    # 첨부파일 복사
                    attachments_dir = item_dir / "첨부파일"
                    if item['attachments']:
                        attachments_dir.mkdir(exist_ok=True)
                        for att in item['attachments']:
                            source_path = self.base_dir / att['path']
                            if source_path.exists():
                                dest_path = attachments_dir / att['name']
                                shutil.copy2(source_path, dest_path)

                elif item['type'] == 'ADDITIONAL_EVIDENCE':
                    f.write("=== 추가 증거 정보 ===\n")
                    f.write(f"증거번호: {item['evidence_number']}\n")
                    f.write(f"카테고리: {item['category']}\n")
                    f.write(f"당사자: {item['party']}\n")
                    f.write(f"파일 해시: {item['file_info']['hash']}\n")

                    if item['related_emails']:
                        f.write(
                            f"관련 이메일: {', '.join(item['related_emails'])}\n")

                    # 실제 파일 복사
                    source_path = self.base_dir / item['file_info']['path']
                    if source_path.exists():
                        dest_path = item_dir / item['file_info']['name']
                        shutil.copy2(source_path, dest_path)
                        f.write(f"\n첨부된 파일: {item['file_info']['name']}\n")

        print(f"  📁 패키지 생성 완료: {package_dir.name}")
        return str(package_dir)

    def _generate_excel_timeline(self, timeline_items: List[Dict]) -> Path:
        """Excel 타임라인 생성"""
        print("📊 Excel 타임라인 생성 중...")

        # DataFrame 생성을 위한 데이터 준비
        excel_data = []

        for i, item in enumerate(timeline_items, 1):
            row = {
                '순번': i,
                '날짜': str(item['event_date']),
                '시간': item['event_time'],
                '유형': '이메일' if item['type'] == 'EMAIL' else '추가증거',
                '제목': item['title'],
                '설명': item['description'][:100] + '...' if len(item['description']) > 100 else item['description'],
                '중요도': item['importance']
            }

            if item['type'] == 'EMAIL':
                row['발신자'] = item['participants']['from']
                row['수신자'] = item['participants']['to']
                row['첨부파일수'] = len(item['attachments'])
                row['증거파일수'] = len(item['evidence_files'])
            else:
                row['증거번호'] = item['evidence_number']
                row['카테고리'] = item['category']
                row['당사자'] = item['party']
                row['파일크기'] = f"{item['file_info']['size'] / 1024:.1f} KB"

            excel_data.append(row)

        # openpyxl로 Excel 파일 생성
        try:
            import openpyxl
        except ImportError:
            print("  ⚠️  openpyxl이 설치되지 않음. Excel 생성을 건너뜁니다.")
            # 임시 경로 반환
            return self.timeline_dir / "timeline_excel_not_available.txt"

        excel_path = self.timeline_dir / \
            f"통합타임라인_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        workbook = openpyxl.Workbook()

        # 메인 타임라인 시트 (기본 시트 사용)
        ws_main = workbook.active
        if ws_main is not None:
            ws_main.title = '통합타임라인'

        # 헤더 정의
        headers = []
        if excel_data and ws_main is not None:
            headers = list(excel_data[0].keys())
            ws_main.append(headers)

            # 데이터 작성
            for row_data in excel_data:
                ws_main.append(list(row_data.values()))

        # 이메일만 필터링된 시트
        email_data = [row for row in excel_data if row['유형'] == '이메일']
        if email_data and headers:
            ws_email = workbook.create_sheet('이메일타임라인')
            ws_email.append(headers)
            for row_data in email_data:
                ws_email.append(list(row_data.values()))

        # 추가증거만 필터링된 시트
        evidence_data = [row for row in excel_data if row['유형'] == '추가증거']
        if evidence_data and headers:
            ws_evidence = workbook.create_sheet('추가증거타임라인')
            ws_evidence.append(headers)
            for row_data in evidence_data:
                ws_evidence.append(list(row_data.values()))

        workbook.save(excel_path)

        print(f"  📊 Excel 파일 생성 완료: {excel_path.name}")
        return excel_path


# 사용 예시
if __name__ == "__main__":
    generator = IntegratedTimelineGenerator()
    result = generator.generate_integrated_timeline()

    print(f"\n📋 타임라인 생성 결과:")
    print(f"  총 항목: {result['total_items']}개")
    print(f"  이메일: {result['email_items']}개")
    print(f"  추가증거: {result['additional_items']}개")
    print(f"  Excel 파일: {result['excel_path']}")
