# src/mail_parser/reporter.py

import os
from datetime import datetime
from typing import Any, Dict, List

# Delay importing openpyxl until actually generating Excel to avoid import-time
# heavy dependency loading (numpy/openpyxl internals). Imports are placed inside
# functions that require them.


class ReportGenerator:
    """
    증거목록 및 처리 보고서 생성 클래스
    """

    def __init__(self, output_dir: str = "processed_emails"):
        self.output_dir = output_dir
        self.evidence_list: List[Dict[str, Any]] = []
        self.processing_stats = {
            'total_emails': 0,
            'processed_emails': 0,
            'excluded_emails': 0,
            'generated_pdfs': 0,
            'total_attachments': 0
        }

    def add_evidence_entry(self, entry: Dict[str, Any]):
        entry['status'] = entry.get('status', '')
        self.evidence_list.append(entry)
        self.processing_stats['processed_emails'] += 1

    def add_excluded_entry(self, entry: Dict[str, Any]):
        entry['status'] = '제외됨'
        self.evidence_list.append(entry)
        self.processing_stats['excluded_emails'] += 1

    def update_stats(self, stat_name: str, value: int = 1):
        if stat_name in self.processing_stats:
            self.processing_stats[stat_name] += value

    def generate_excel_report(self, output_path: str = None, party: str = "갑") -> str:
        # prepare output path
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(
                self.output_dir, f"증거목록_{party}_{timestamp}.xlsx")

        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        # lazy import openpyxl to avoid heavy import on module load
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        self._create_evidence_sheet(wb, party)
        self._create_statistics_sheet(wb)
        self._create_excluded_sheet(wb)

        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        try:
            wb.save(output_path)
            return output_path
        except Exception as e:
            raise Exception(f"Excel 파일 저장 실패: {str(e)}")

    def _create_evidence_sheet(self, wb, party: str):
        # lazy imports
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        ws = wb.active
        ws.title = f"{party} 증거목록"

        headers = [
            "번호", "증거번호", "날짜", "제목", "발신자", "수신자",
            "첨부파일", "파일경로", "해시값(SHA-256)", "상태", "비고"
        ]

        header_font = Font(name='맑은 고딕', size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        processed_entries = [
            entry for entry in self.evidence_list if entry.get('status') != '제외됨']

        for row, entry in enumerate(processed_entries, 2):
            ws.cell(row=row, column=1, value=row-1)
            ws.cell(row=row, column=2, value=entry.get('evidence_number', ''))
            ws.cell(row=row, column=3, value=entry.get('date', ''))
            ws.cell(row=row, column=4, value=entry.get('subject', ''))
            ws.cell(row=row, column=5, value=entry.get('sender', ''))
            ws.cell(row=row, column=6, value=entry.get('receiver', ''))

            attachments = entry.get('attachments', [])
            attachment_str = ", ".join(attachments) if attachments else "없음"
            ws.cell(row=row, column=7, value=attachment_str)

            ws.cell(row=row, column=8, value=entry.get('file_path', ''))
            ws.cell(row=row, column=9, value=entry.get('hash_value', ''))
            ws.cell(row=row, column=10, value=entry.get('status', ''))
            ws.cell(row=row, column=11, value=entry.get('notes', ''))

        column_widths = [5, 15, 12, 40, 25, 25, 30, 50, 70, 10, 20]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        thin_border = Border(left=Side(style='thin'), right=Side(
            style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for row in ws.iter_rows(min_row=1, max_row=len(processed_entries)+1, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
                if cell.row > 1:
                    cell.font = Font(name='맑은 고딕', size=10)
                    cell.alignment = Alignment(
                        vertical="center", wrap_text=True)

    def _create_statistics_sheet(self, wb):
        from openpyxl.styles import Font
        ws = wb.create_sheet("처리 통계")

        ws.cell(row=1, column=1, value="메일박스 처리 통계").font = Font(
            size=14, bold=True)
        ws.cell(row=2, column=1,
                value=f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

        stats_data = [
            ("구분", "개수"),
            ("전체 메일 수", self.processing_stats['total_emails']),
            ("처리된 메일 수", self.processing_stats['processed_emails']),
            ("제외된 메일 수", self.processing_stats['excluded_emails']),
            ("생성된 PDF 수", self.processing_stats['generated_pdfs']),
            ("총 첨부파일 수", self.processing_stats['total_attachments'])
        ]

        for row, (label, value) in enumerate(stats_data, 4):
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)

            if row == 4:
                ws.cell(row=row, column=1).font = Font(bold=True)
                ws.cell(row=row, column=2).font = Font(bold=True)

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15

    def _create_excluded_sheet(self, wb):
        from openpyxl.styles import Font, PatternFill

        ws = wb.create_sheet("제외된 메일")

        headers = ["번호", "날짜", "제목", "발신자", "제외 사유"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

        excluded_entries = [
            entry for entry in self.evidence_list if entry.get('status') == '제외됨']

        for row, entry in enumerate(excluded_entries, 2):
            ws.cell(row=row, column=1, value=row-1)
            ws.cell(row=row, column=2, value=entry.get('date', ''))
            ws.cell(row=row, column=3, value=entry.get('subject', ''))
            ws.cell(row=row, column=4, value=entry.get('sender', ''))
            ws.cell(row=row, column=5, value=entry.get('exclusion_reason', ''))

        column_widths = [5, 12, 40, 25, 30]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

    def generate_summary_report(self, output_path: str = None) -> str:
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(
                self.output_dir, f"처리요약_{timestamp}.txt")

        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        with open(output_path, 'w', encoding='utf-8') as f:
            f.write("="*50 + "\n")
            f.write("메일박스 증거 처리 요약 보고서\n")
            f.write("="*50 + "\n")
            f.write(
                f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")

            f.write("처리 통계:\n")
            f.write("-"*30 + "\n")
            for key, value in self.processing_stats.items():
                korean_labels = {
                    'total_emails': '전체 메일 수',
                    'processed_emails': '처리된 메일 수',
                    'excluded_emails': '제외된 메일 수',
                    'generated_pdfs': '생성된 PDF 수',
                    'total_attachments': '총 첨부파일 수'
                }
                label = korean_labels.get(key, key)
                f.write(f"{label}: {value}개\n")

            f.write(
                f"\n처리율: {self.processing_stats['processed_emails']/max(self.processing_stats['total_emails'], 1)*100:.1f}%\n")

            if self.processing_stats['excluded_emails'] > 0:
                f.write(f"\n제외된 메일의 주요 사유:\n")
                f.write("-"*30 + "\n")

        return output_path
